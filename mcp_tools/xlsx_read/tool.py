"""Read data from an Excel XLSX file using openpyxl."""

import json
import os


def run(file_path: str, sheet: str = "") -> str:
    """
    Read all data from an Excel file and return it as a JSON string.

    @param file_path: Path to the Excel file to read.
    @param sheet: Name of the sheet to read. Uses the active sheet if empty.
    @returns: JSON string containing an array of rows (each row is an array of cell values).
    @throws FileNotFoundError: If the Excel file does not exist.
    @throws ImportError: If openpyxl is not installed.
    """
    try:
        import openpyxl
    except ImportError:
        return "error: " + "openpyxl is required but not installed. "
            "Install it with: pip install openpyxl"

    resolved = os.path.abspath(file_path)
    if not os.path.exists(resolved):
        raise FileNotFoundError(f"Excel file not found: {resolved}")

    workbook = openpyxl.load_workbook(resolved, read_only=True, data_only=True)

    if sheet:
        if sheet not in workbook.sheetnames:
            available = ", ".join(workbook.sheetnames)
            workbook.close()
            raise ValueError(
                f"Sheet '{sheet}' not found. Available sheets: {available}"
            )
        worksheet = workbook[sheet]
    else:
        worksheet = workbook.active

    rows: list[list[object]] = []
    for row in worksheet.iter_rows(values_only=True):
        rows.append([_serialize_cell(cell) for cell in row])

    workbook.close()

    if not rows:
        return "No data found in the spreadsheet."

    return json.dumps(rows, indent=2, default=str)


def _serialize_cell(value: object) -> object:
    """Convert a cell value to a JSON-safe type."""
    if value is None:
        return None
    if isinstance(value, (int, float, bool, str)):
        return value
    return str(value)
